import re
from datetime import datetime

import pandas as pd
from fastapi import UploadFile
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.entities.report import ReportInDTO, ReportFilters, ReportInfoOutDTO, Metric


class GetReportInfoService:

    def execute(self, report_in_dto: ReportInDTO) -> ReportInfoOutDTO:
        self.__raise_if_file_is_invalid(report_in_dto.file)
        workbook = self.process_xlsx(report_in_dto.file)
        report_info = self.get_metrics(workbook, report_in_dto.filters)
        return report_info

    @staticmethod
    def __raise_if_file_is_invalid(file: UploadFile):
        if file.headers['content-type'] != 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            raise Exception

    def process_xlsx(self, file: UploadFile) -> Workbook:
        file.file.seek(0)
        workbook = load_workbook(file.file, data_only=True)
        workbook_processed = Workbook()
        accepted_sheets = ['Por Município', 'Profissionais ativos_PA', 'Profissionais ativos_PB']

        if 'Sheet' in workbook_processed.sheetnames:
            std = workbook_processed['Sheet']
            workbook_processed.remove(std)

        for sheet_name in workbook.sheetnames:
            if sheet_name in accepted_sheets:
                sheet = workbook[sheet_name]
                sheet_processed = self.delete_merged_rows(sheet)
                new_sheet = workbook_processed.create_sheet(title=sheet_name)
                for row in sheet_processed.iter_rows(values_only=True):
                    new_sheet.append(row)
        return workbook_processed

    def delete_merged_rows(self, sheet: Worksheet) -> Worksheet:
        merged_rows = self.get_merged_rows(sheet)
        for merge_range in list(sheet.merged_cells.ranges):
            sheet.unmerge_cells(str(merge_range))
        for row in merged_rows:
            sheet.delete_rows(row)
        return sheet

    @staticmethod
    def get_merged_rows(sheet):
        merged_cells_str = str([str(merged_cell_str) for merged_cell_str in sheet.merged_cells.ranges])
        merged_lines_numbers = set(re.findall(r'\d+', merged_cells_str))
        numbers_list = [int(merged_lines_number) for merged_lines_number in merged_lines_numbers]
        numbers_list.sort(reverse=True)
        return numbers_list

    @staticmethod
    def get_dataframe(workbook: Workbook, df_type: str):
        if df_type == 'REGIONAL':
            sheet = workbook.get_sheet_by_name('Por Município')
            data = sheet.values
            columns = next(data)
            df = pd.DataFrame(data, columns=columns)
            return df
        elif df_type == 'PROFISSIONAL':
            df = pd.DataFrame()
            for sheet_name in workbook.sheetnames:
                if 'profissionais' in sheet_name.lower():
                    sheet = workbook.get_sheet_by_name(sheet_name)
                    data = sheet.values
                    columns = next(data)
                    df_sheet = pd.DataFrame(data, columns=columns)
                    df_sheet["Categoria"] = sheet_name
                    df = pd.concat([df, df_sheet])
            return df
        return None

    def get_metrics(self, workbook: Workbook, filters: ReportFilters) -> ReportInfoOutDTO:
        df_regiao = self.get_dataframe(workbook, 'REGIONAL')
        df_profissionais = self.get_dataframe(workbook, 'PROFISSIONAL')
        if filters.type == 'REGIONAL':
            if filters.scope == 'REGIÃO':
                return ReportInfoOutDTO(
                    title=f'REGIÃO - {filters.value.upper()}',
                    metrics=self.get_metrics_regional_regiao(df_regiao, df_profissionais, filters.value.upper())
                )
            elif filters.scope == 'UF':
                return ReportInfoOutDTO(
                    title=f'UF - {filters.value.upper()}',
                    metrics=self.get_metrics_regional_uf(df_regiao, df_profissionais, filters.value.upper())
                )
            elif filters.scope == 'MUNICÍPIO':
                return ReportInfoOutDTO(
                    title=f'MUNICÍPIO - {filters.value.upper()}',
                    metrics=self.get_metrics_regional_municipio(df_regiao, df_profissionais, filters.value.upper())
                )
            else:
                pass
        else:
            return ReportInfoOutDTO(
                title=f'PROFISSIONAL',
                metrics=self.get_metrics_profissional(df_profissionais, filters.value)
            )


    @staticmethod
    def get_metrics_regional_regiao(df_regiao: pd.DataFrame, df_profissionais: pd.DataFrame, regiao: str) -> list[Metric]:
        municipios_regiao = df_regiao[df_regiao['Região'].str.upper() == regiao]['Município'].unique()
        quantidade_estados = df_regiao[df_regiao['Região'].str.upper() == regiao]['UF'].nunique()
        quantidade_populacao = df_regiao[df_regiao['Região'].str.upper() == regiao]['População 2021'].sum()
        quantidade_profissionais = df_profissionais[df_profissionais['Municipio/DEEI'].isin(municipios_regiao)][
            'CPF'].nunique()
        metrics = [
            Metric(metric="Quantidade de estados", value=quantidade_estados),
            Metric(metric="População total", value=quantidade_populacao),
            Metric(metric="quantidade de profissionais", value=quantidade_profissionais)
        ]
        return metrics

    @staticmethod
    def get_metrics_regional_uf(df_regiao: pd.DataFrame, df_profissionais: pd.DataFrame, uf: str) -> list[Metric]:
        municipios_uf = df_regiao[df_regiao['UF'] == uf]['Município'].unique()
        df_regiao_uf = df_regiao[df_regiao['UF'] == uf]
        df_profissionais_uf = df_profissionais[df_profissionais['Municipio/DEEI'].isin(municipios_uf)]
        quantidade_populacao = df_regiao_uf['População 2021'].sum()
        potencial_cobertura = df_regiao_uf['Potencial de cobertura da população pelo Programa '].sum()
        quantidade_municipios = df_regiao_uf['Município'].nunique()
        municipios_contemplados = df_regiao_uf[df_regiao_uf['Total de vagas ocupadas'] > 0]['Município'].nunique()
        percentual_contemplados = round((municipios_contemplados / quantidade_municipios) * 100,
                                        2) if quantidade_municipios > 0 else 0
        quantidade_profissionais = df_profissionais_uf['CPF'].nunique()
        metrics = [
            Metric(metric="População total", value=quantidade_populacao),
            Metric(metric="Total de profissionais", value=quantidade_profissionais),
            Metric(metric="Potencial de cobertura do programa", value=potencial_cobertura),# REVISAR
            Metric(metric="Total de municípios", value=quantidade_municipios),
            Metric(metric="Total de municípios contemplados", value=municipios_contemplados),# REVISAR
            Metric(metric="Municípios contemplados (%)", value=f'{percentual_contemplados}%'),# REVISAR
        ]
        return metrics

    @staticmethod
    def get_metrics_regional_municipio(df_regiao: pd.DataFrame, df_profissionais: pd.DataFrame, municipio: str) -> list[
        Metric]:
        df_regiao_mun = df_regiao[df_regiao['Município'] == municipio]
        df_profissionais_mun = df_profissionais[df_profissionais['Municipio/DEEI'] == municipio]
        quantidade_populacao = df_regiao_mun['População 2021'].sum()
        potencial_cobertura = df_regiao_mun['Potencial de cobertura da população pelo Programa '].sum()
        quantidade_profissionais = df_profissionais_mun['CPF'].nunique()
        indice_vulnerabilidade = df_regiao_mun['Categoria de IVS'].values[
            0] if not df_regiao_mun.empty else "Não informado"
        metrics = [
            Metric(metric="População total", value=quantidade_populacao),
            Metric(metric="Total de profissionais", value=quantidade_profissionais),
            Metric(metric="Potencial de cobertura do programa", value=potencial_cobertura),
            Metric(metric="Índice de vulnerabilidade social", value=indice_vulnerabilidade)
        ]
        return metrics

    @staticmethod
    def get_metrics_profissional(df_profissionais: pd.DataFrame, cpf: str) -> list[Metric]:
        def format_cpf(cpf):
            cleaned_cpf = re.sub(r'\D', '', str(cpf))
            if len(cleaned_cpf) == 11:
                return f"XXX.{cleaned_cpf[3:6]}.XXX-{cleaned_cpf[9:]}"
            return cpf

        quantidade_total_profissionais = df_profissionais['CPF'].nunique()
        quantidade_pa = df_profissionais[df_profissionais['Categoria'] == 'Profissionais ativos_PA']['CPF'].nunique()
        quantidade_pb = df_profissionais[df_profissionais['Categoria'] == 'Profissionais ativos_PB']['CPF'].nunique()
        df_profissionais['cpf_limpo'] = df_profissionais['CPF'].str.replace(r'\D', '', regex=True)
        profissional = df_profissionais[df_profissionais['cpf_limpo'] == re.sub(r'\D', '', str(cpf))]
        if profissional.empty:
            return []
        profissional = profissional.iloc[0]
        metrics = [
            Metric(metric="[GERAL] Total de profissionais", value=quantidade_total_profissionais),
            Metric(metric="[GERAL] Profissionais PA", value=quantidade_pa),
            Metric(metric="[GERAL] Profissionais PB", value=quantidade_pb),
            Metric(metric="CPF", value=format_cpf(profissional['CPF'])),
            Metric(metric="Nome completo", value=profissional['Nome']),
            Metric(metric="Ciclo", value=profissional['Ciclo']),
            Metric(metric="Perfil", value=profissional['Perfil do Profissional']), #ADD CONDICIONAL CODIGO B
            Metric(metric="Sexo", value=profissional['Sexo']),
            Metric(metric="Idade", value='N/A'),
            Metric(metric="Raça/cor", value=profissional['Raça/Cor']),
            Metric(metric="Nacionalidade", value=profissional['Nacionalidade']),
            Metric(metric="Município", value=profissional['Municipio/DEEI'])
        ]
        return metrics


